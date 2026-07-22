"""Excel 输出模块：将采集结果写入xlsx文件"""

from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .utils import logger

# 结果行类型
# (content, email, has_content)
# has_content: "有内容" 或 "主页无内容"
ResultRow = Tuple[str, str, str]

# 单元格样式
_cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
_data_font = Font(name="等线", size=11)
_header_font = Font(name="等线", bold=True, size=11)
_row_height = 30  # 行高（磅）


def _create_workbook() -> Workbook:
    """创建工作簿并设置列宽"""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "教师简介"

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16

    return wb


def write_excel(results: List[ResultRow], filename: str) -> str:
    """将结果列表写入 Excel 文件"""
    wb = _create_workbook()
    ws = wb.active
    assert ws is not None

    for row_idx, (content, email, has_content) in enumerate(results, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=content)
        cell_a.font = _data_font
        cell_a.alignment = _cell_align

        cell_b = ws.cell(row=row_idx, column=2, value=email)
        cell_b.font = _data_font
        cell_b.alignment = _cell_align

        cell_c = ws.cell(row=row_idx, column=3, value=has_content)
        cell_c.font = _data_font
        cell_c.alignment = _cell_align

        # 设置数据行行高
        ws.row_dimensions[row_idx].height = _row_height

    wb.save(filename)
    logger.info(f"输出文件: {filename}")
    return filename
